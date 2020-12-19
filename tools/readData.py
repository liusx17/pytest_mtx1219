import openpyxl
import yaml

from setting import DIR_NAME

class ReadData:
    '''
    列表嵌套列表，作业：列表嵌套字典
    # data_li = [
#     ['yaoyao','yaoyao','登录成功'],
#     ['lwx','yaoyao','帐号不存在'],
#     ['yaoyao','123456','密码错误1111111']
# ]
    '''
    # def get_excel(self):
    #     # wb 工作目录
    #     wb = openpyxl.load_workbook('./data.xlsx')
    #     # ws  sheet页
    #     ws = wb['测试用例']
    #     # 构造数据  [[],[],[]]
    #     all_cases=[]
    #     selected_data_area = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)
    #     # print(selected_data_area)
    #     for rows in selected_data_area:
    #         cell_value_list = [cell.value for cell in rows]
    #         # print('每行的cell值的列表是{}'.format(cell_value_list))
    #         all_cases.append(cell_value_list)
    #
    #     return all_cases

    # def get_excel(self, header):
    #     '''
    #     [{},{},{}]
    #     all_case = []
    #     dict = {}  # 键--固定好的，用户自定义，以参数的形式进行传递
    #     值--> 读取excel：book-sheet-cell-value
    #     [{'username':'yaoyao','password':'yaoyao','exp':'登录成功'}，...]
    #     all_case.append(dict)
    #     header: []列表，每一个元素就是自己定义的头信息，即关键字,e.g. ['username','password','exp]
    #     :return:
    #     '''
    #     # excel的文档
    #     wb = openpyxl.load_workbook('data.xlsx')
    #     # ws
    #     ws = wb['测试用例']
    #     # 列表容器
    #     all_cases = []
    #     # 列数  是header里面元素的个数是一致的
    #     # 父循环控制行数(2,3,4) 思考:构造的数据，还是要循环的次数
    #     for row in range(2,ws.max_row+1):
    #         # 构造空字典,存放每一行的数据
    #         case_dict = {}
    #         # 子循环控制列数 1.主要循环次数，就可以确定现在这次循环是处于第几列2，也要数据，key，构造字典
    #         # 自定义控制列数,初始值
    #         col =1
    #         for key in header:
    #             # 构造字典
    #             case_dict[key]=ws.cell(row,col).value
    #             # 增量
    #             col += 1
    #         # 子循环结束，一行里面的数据已经构造好了
    #         all_cases.append(case_dict)
    #
    #     return all_cases


    def get_yaml(self,filename,key):
        '''
        解析yaml，yml文件，得到列表嵌套字典的数据格式
        ===============
        以下是读取json的命令
        json.load()  读
        json.dump()  写
        ===================
        以下是读取yaml的命令
        yaml.safe_load()
        :return:
        # data_li = [
#     {'accounts': 'yaoyao', 'pwd':'yaoyao','exp':'登录成功'},
#     {'accounts': 'lwx', 'pwd': 'yaoyao','exp':'帐号不存在'},
#
# ]
        '''
        with open(DIR_NAME + '/data/%s.yml' % filename, encoding='utf-8') as f:
            # 最新的知识点
            yaml_data = yaml.safe_load(f)
            # print(yaml_data)
            # 构造空的列表
            data_list = []
            cases_dict = yaml_data.get(key)
            # case_object是可迭代对象  for ...in..,list(),列表.extend()
            case_object = cases_dict.values()
            # print(case_object)
            # data_list.extend(case_object)
            data = list(case_object)
            return data

#     def get_yaml1(self, key):
#         '''
#         解析yaml，yml文件，得到列表嵌套字典的数据格式
#         ===============
#         以下是读取json的命令
#         json.load()  读
#         json.dump()  写
#         ===================
#         以下是读取yaml的命令
#         yaml.safe_load()
#         :return:
#         # data_li = [
# #     {'accounts': 'yaoyao', 'pwd':'yaoyao','exp':'登录成功'},
# #     {'accounts': 'lwx', 'pwd': 'yaoyao','exp':'帐号不存在'},
# #
# # ]
#         '''
#         with open('./data/login_data.yml', encoding='utf-8') as f:
#             # 最新的知识点
#             yaml_data = yaml.safe_load(f)
#             # print(yaml_data)
#             # 构造空的列表
#             data_list = []
#             cases_dict = yaml_data.get(key)
#             # case_object是可迭代对象  for ...in..,list(),列表.extend()
#             case_object = cases_dict.values()
#             # print(case_object)
#             # data_list.extend(case_object)
#             data = list(case_object)
#             return data


if __name__ == '__main__':
    # all_cases=ReadExcel().get_excel()
    # print(all_cases)
    print(ReadData().get_yaml('login_data', 'test_login'))
    # result = ReadData().get_excel(['username', 'password', 'exp'])
    # print(result)

